import urllib.request
from urllib import parse
import json
from openpyxl import load_workbook
nameList = []
book = load_workbook('/eclipse-workspace/excel/driving.xlsx')
nameSheet = book['Sheet1']
for row in range(1,nameSheet.max_row+1):
    nameList.append(str(nameSheet['A%d'%(row)].value))
dict = {}
for i in nameList:
    url1 = 'http://restapi.amap.com/v3/place/text?keywords=%E6%A6%95%E5%9F%8E&city=445202&output=json&offset=1&page=1&key=9f5d863e1b5694dc5e37ece0ab5d19b8'
    newUrl1 = parse.quote(url1, safe="/:=&?#+!$,;'@()*[]")
    response1 = urllib.request.urlopen(newUrl1)
    data1 = response1.read()
    jsonData1 = json.loads(data1)
    dict[i] = jsonData1['pois'][0]['location']
    locations = dict[i].split(',')
distanceList = []
k = len(nameList)
for m in range(k):
    subList = []
    for n in range(k):
        origin = dict[nameList[m]]
        destination = dict[nameList[n]]
        url2 = 'https://restapi.amap.com/v3/direction/driving?origin=116.367035,23.525229&destination=113.361200,23.124680&extensions=all&strategy=2&output=json&key=412b5b466b09174dbb9b8783492a75f4'
        newUrl2 = parse.quote(url2, safe="/:=&?#+!$,;'@()*[]")
        response2 = urllib.request.urlopen(newUrl2)
        data2 = response2.read()
        jsonData2 = json.loads(data2)
        distance = jsonData2['route']['paths'][0]['distance']
        subList.append(int(distance))
        print(nameList[m],nameList[n],distance)
    distanceList.append(subList)